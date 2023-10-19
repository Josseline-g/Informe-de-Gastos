import openpyxl

# Cargar el archivo Excel
archivo_excel = "./informe_gastos.xlsx"
libro_trabajo = openpyxl.load_workbook(archivo_excel)

# Acceder a una hoja 
hoja = libro_trabajo["Gastos"]


# Insertar una nueva fila después de la fila 
hoja.insert_rows(2)

# Insertar un nuevo valor en la celda 
print("Inserte fecha")
hoja['A2'] = fecha = input()

print("Inserte Gasto")
hoja['B2'] = gasto = input()

print("Inserte Monto")
hoja['C2'] = monto = input()


# Recorrer las filas de la hoja de cálculo
for fila in hoja.iter_rows(min_row=1, values_only=True):
    print(fila)
    
# Recorrer las columnas de la hoja de cálculo
for columna in hoja.iter_cols (values_only=True):
    print(columna)

# Recorrer las celdas de la hoja de cálculo
for fila in hoja.iter_rows(min_row=1, values_only=True):
    for celda in fila:
        print(celda)

# Guardar el archivo Excel modificado
libro_trabajo.save(archivo_excel)

# Cerrar el archivo Excel
libro_trabajo.close()
